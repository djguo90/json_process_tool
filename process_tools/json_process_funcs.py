from typing import Any, List, Optional, Union
from dataclasses import dataclass
import glob
import json
from pathlib import Path
import openpyxl
from openpyxl.styles import Font
import tqdm
import string
"""
1. 删除字段
2. 重命名字段
3. 内部根据字段去重
4. 外部根据字段去重
"""
@dataclass
class PathStep:
    key: str
    mode: str  # "dict" 普通字段, "all" 列表所有元素 key[], "indices" 指定下标 key[0,1]
    indices: Optional[List[int]] = None


def _parse_path(path: str) -> List[PathStep]:
    """
    将路径字符串解析为 PathStep 列表。
    支持：
      .key1.key2.key3[].key4
      .key1.key2.key3[0,1].key4
      .key1.key2.key3[-1,-2].key4
    """
    path = path.strip()
    # 去掉前导的点
    while path.startswith("."):
        path = path[1:]

    if not path:
        return []

    raw_parts = [p.strip() for p in path.split(".") if p.strip()]

    steps: List[PathStep] = []
    for part in raw_parts:
        part = part.strip()
        # 通配所有元素 key[]
        if part.endswith("[]"):
            key = part[:-2].strip()
            steps.append(PathStep(key=key, mode="all"))
            continue

        # 指定下标 key[0,1,-1]
        if "[" in part and part.endswith("]"):
            key, idx_part = part.split("[", 1)
            key = key.strip()
            idx_str = idx_part[:-1].strip()  # 去掉 ']'
            indices: List[int] = []
            if idx_str:
                for tok in idx_str.split(","):
                    tok = tok.strip()
                    if not tok:
                        continue
                    try:
                        # 这里直接 int()，支持负数
                        indices.append(int(tok))
                    except ValueError:
                        # 忽略非法下标（也可以选择抛异常）
                        pass
            steps.append(PathStep(key=key, mode="indices", indices=indices))
            continue

        # 普通 key
        steps.append(PathStep(key=part, mode="dict"))

    return steps


def delete_field_by_path(data: Any, path: str) -> None:
    """
    根据类似 .key1.key2.key3[].key4 / .key1.key2.key3[0,1].key4 / .key1.key2.key3[-1,-2].key4 的路径，
    在 data 中删除对应字段。原地修改（in-place），无返回值。

    特殊说明（最后一段的语义）：
    - ...keyX            ：删除当前对象中的字段 keyX
    - ...keyX[]          ：清空keyX对应的列表，不删除该字段
    - ...keyX[0,1,-1]    ：从 keyX 列表中删除这些下标对应的元素（支持负下标）
    """
    steps = _parse_path(path)
    if not steps:
        return

    def _delete(cur: Any, idx: int) -> None:
        if idx >= len(steps):
            return

        step = steps[idx]
        is_last = (idx == len(steps) - 1)

        # 普通字典字段 key
        if step.mode == "dict":
            if not isinstance(cur, dict):
                return
            if step.key not in cur:
                return

            if is_last:
                # 最后一段：直接删字段
                cur.pop(step.key, None)
                return

            nxt = cur.get(step.key)

            # 继续往下走
            if isinstance(nxt, dict):
                _delete(nxt, idx + 1)
            elif isinstance(nxt, list):
                # 如果下一级是列表，默认对列表里每个元素继续走
                for item in nxt:
                    _delete(item, idx + 1)
            else:
                # 原子类型，没法再深入
                return

        # key[]：对列表中所有元素生效
        elif step.mode == "all":
            if not isinstance(cur, dict):
                return
            if step.key not in cur:
                return

            lst = cur.get(step.key)
            if not isinstance(lst, list):
                return

            if is_last:
                # 最后一段：删除整个 key 这个字段
                # cur.pop(step.key, None)
                cur[step.key] = []
                return

            # 对列表中每个元素继续往下
            for item in lst:
                _delete(item, idx + 1)

        # key[indices]：只对列表中指定下标生效（支持负下标）
        elif step.mode == "indices":
            if not isinstance(cur, dict):
                return
            if step.key not in cur:
                return

            lst = cur.get(step.key)
            if not isinstance(lst, list):
                return

            raw_indices = step.indices or []

            # 先把负下标转成真实下标
            real_indices: List[int] = []
            n = len(lst)
            for i in raw_indices:
                if i < 0:
                    j = n + i   # Python 负下标语义
                else:
                    j = i
                if 0 <= j < n:
                    real_indices.append(j)

            if not real_indices:
                return

            if is_last:
                # 最后一段：从列表中删除这些位置的元素
                # 要去重 & 从大到小删，防止位置移动
                unique_valid_indices = sorted(set(real_indices), reverse=True)
                for i in unique_valid_indices:
                    del lst[i]
                return

            # 非最后一段：仅对这些位置的元素继续往下（此处不会改动列表长度）
            for i in set(real_indices):
                _delete(lst[i], idx + 1)

    _delete(data, 0)


def delete_fields(samples, paths: Union[str, List[str]]) -> Any:
    """
    支持一次删多个路径：
    - paths 可以是一个字符串路径
    - 也可以是路径字符串列表
    函数会直接修改 samples 上的每个元素，并yield sample 方便链式使用。
    """
    if isinstance(paths, str):
        paths = [paths]
    for sample in samples:
        for p in paths:
            delete_field_by_path(sample, p)
        yield sample


def rename_field_by_path(data: Any, path: str, new_key: str) -> None:
    """
    根据路径把“末级字段名”改成 new_key。

    示例：
        .key1.key2.key3          -> 把 key3 改名为 new_key
        .key1.key2.lst[].oldKey  -> 把 lst 列表中每个元素里的 oldKey 改名为 new_key
        .key1.lst[0,-1].oldKey   -> 只对第 0 个和最后一个元素里的 oldKey 改名

    约束：
        - 路径最后一段必须是普通 key（不能带 [] 或下标）。
    """
    steps = _parse_path(path)
    if not steps:
        return

    # 最后一段必须是“普通字段”
    last = steps[-1]
    if last.mode != "dict":
        raise ValueError(f"重命名路径最后一段必须是普通字段名，例如 '.a.b.c'，而不是 '{last}'")

    # 如果新旧字段名相同，直接返回
    if new_key == last.key:
        return

    def _rename(cur: Any, idx: int) -> None:
        if idx >= len(steps):
            return

        step = steps[idx]
        is_last = (idx == len(steps) - 1)

        # 普通字典字段
        if step.mode == "dict":
            if not isinstance(cur, dict):
                return
            if step.key not in cur:
                return

            if is_last:
                # 在当前 dict 中把 step.key 改名为 new_key
                value = cur.pop(step.key)
                # 如果 new_key 已存在，这里默认覆盖（如需保留，可自己改策略）
                cur[new_key] = value
                return

            nxt = cur.get(step.key)

            if isinstance(nxt, dict):
                _rename(nxt, idx + 1)
            elif isinstance(nxt, list):
                for item in nxt:
                    _rename(item, idx + 1)
            else:
                return

        # key[]：对列表所有元素继续往下找，直到最后一段再改名
        elif step.mode == "all":
            if not isinstance(cur, dict):
                return
            if step.key not in cur:
                return

            lst = cur.get(step.key)
            if not isinstance(lst, list):
                return

            for item in lst:
                _rename(item, idx + 1)

        # key[indices]：只对指定下标的元素继续往下
        elif step.mode == "indices":
            if not isinstance(cur, dict):
                return
            if step.key not in cur:
                return

            lst = cur.get(step.key)
            if not isinstance(lst, list):
                return

            raw_indices = step.indices or []
            n = len(lst)
            real_indices: List[int] = []
            for i in raw_indices:
                j = n + i if i < 0 else i  # 负下标转正下标
                if 0 <= j < n:
                    real_indices.append(j)

            for j in set(real_indices):
                _rename(lst[j], idx + 1)

    _rename(data, 0)


def rename_fields(samples, mapping: dict[str, str]):
    """
    一次性重命名多个字段。
    mapping: { path: new_key }
    例如：
        {
          ".key1.key2.key3": "key3-1",
          ".key1.key2.lst[].old": "new"
        }
    """
    for sample in samples:
        for path, new_key in mapping.items():
            rename_field_by_path(sample, path, new_key)
        yield sample


def _get_value_by_path_simple(item: Any, path: str) -> Any:
    """
    从 item 中根据路径取出对应值。
    若路径不存在，返回 None。
    """
    parts = [p.strip() for p in path.split(".") if p.strip()]
    cur = item
    for p in parts:
        if isinstance(cur, dict):
            if p in cur:
                cur = cur[p]
            else:
                return None
        else:
            return None
    return cur


# TODO 这个支持简单key_paths，也即不带[]的
def remove_duplicates_interior(samples, key_paths: List[str]):
    """
    去重函数：
    - samples: 一个列表（其中每个元素通常是 dict）
    - key_paths: 主键路径列表（例如 [".id", ".info.id"]）

    返回：去重后的新列表
    """
    if isinstance(key_paths, str):
        key_paths = [key_paths]
    seen = set()
    # result = []
    for sample in samples:
        key_list = []
        for path in key_paths:
            v = _get_value_by_path_simple(sample, path)
            if v is not None:
                key_list.append(str(v))
            else:
                key_list.append(None)
        key_tuple = tuple(key_list)
        if key_tuple not in seen:
            seen.add(key_tuple)
            yield sample


def remove_duplicates_exterior(samples, target_file_patterns, key_paths):
    """
    外部去重
    - samples: 一个列表（其中每个元素通常是 dict）
    - target_file_patterns: 一些file patterns，这些文件中出现过的数据，就过滤掉
    - key_paths: 主键路径列表（例如 [".id", ".info.id"]）

    返回：去重后的新列表
    """
    if isinstance(target_file_patterns, str):
        target_file_patterns = [target_file_patterns]
    if isinstance(key_paths, str):
        key_paths = [key_paths]
    target_file_paths = []
    for tfp in target_file_patterns:
        target_file_paths.extend(glob.glob(tfp, recursive=True))
    target_file_paths = sorted(target_file_paths)
    target_key_set = set()
    for tfp in target_file_paths:
        with open(tfp, "r", encoding="utf-8") as reader:
            for l in reader:
                # try:
                l_json = json.loads(l)
                key_list = []
                for path in key_paths:
                    v = _get_value_by_path_simple(l_json, path)
                    if v is not None:
                        key_list.append(str(v))
                    else:
                        key_list.append(None)
                key_tuple = tuple(key_list)
                target_key_set.add(key_tuple)
    for sample in samples:
        key_list = []
        for path in key_paths:
            v = _get_value_by_path_simple(sample, path)
            if v is not None:
                key_list.append(str(v))
            else:
                key_list.append(None)
        key_tuple = tuple(key_list)
        if key_tuple not in target_key_set:
            yield sample


def has_key_path(item: Any, key_path: str) -> bool:
    """
    判断一条 item 中是否“存在”这个 key_path（支持 [] / [0,-1]）。

    语义约定：

    - .a.b.c      ：普通字典下钻，所有 key 都存在就算存在；
    - .a.b[]      ：a.b 是一个 list，且至少有一个元素 就算存在；
    - .a.b[].c    ：a.b 是 list，且所有元素里都存在路径 .c；
    - .a.b[0,-1].c：a.b 是 list，取下标 0 和倒数第一个元素，这两个元素都满足 .c 路径，就算存在；
    - 负下标：和 Python 一样，-1 表示最后一个。
    """

    steps = _parse_path(key_path)
    if not steps:
        # 空路径直接认为存在
        return True

    def _exists(cur: Any, idx: int) -> bool:
        if idx >= len(steps):
            # 所有片段都走完，说明路径完整存在
            return True

        step = steps[idx]
        is_last = (idx == len(steps) - 1)

        # 1) 普通字典字段：.a.b.c
        if step.mode == "dict":
            if not isinstance(cur, dict):
                return False
            if step.key not in cur:
                return False
            nxt = cur[step.key]
            return _exists(nxt, idx + 1)

        # 2) key[]：列表所有元素
        elif step.mode == "all":
            if not isinstance(cur, dict):
                return False
            if step.key not in cur:
                return False
            lst = cur[step.key]
            if not isinstance(lst, list):
                return False

            if is_last:
                # .a.b[]：是 list 且至少有一个元素
                return len(lst) > 0

            # .a.b[].c：a.b 是 list，且所有元素里都存在路径 .c
            if len(lst) == 0:
                return False  # 空列表不算“所有元素都有”
            for elem in lst:
                if not _exists(elem, idx + 1):
                    return False
            return True

        # 3) key[indices]：列表指定下标
        elif step.mode == "indices":
            if not isinstance(cur, dict):
                return False
            if step.key not in cur:
                return False
            lst = cur[step.key]
            if not isinstance(lst, list):
                return False

            raw_indices = step.indices or []
            if not raw_indices:
                return False  # 没指定任何下标，直接认为不存在

            n = len(lst)
            elems: List[Any] = []

            # 负下标语义：-1 表示最后一个
            for raw_i in raw_indices:
                j = raw_i if raw_i >= 0 else n + raw_i
                if 0 <= j < n:
                    elems.append(lst[j])
                else:
                    # 只要有一个下标越界，就认为路径不存在
                    return False

            if is_last:
                # .a.b[0,-1] 作为最后一段：
                # 只要这些下标都合法（已经检查过）且列表非空，就算存在
                return len(elems) > 0

            # .a.b[0,-1].c：这几个元素都要满足后续路径 .c
            for elem in elems:
                if not _exists(elem, idx + 1):
                    return False
            return True

        else:
            return False

    return _exists(item, 0)


def _resolve_indices(n: int, raw_indices: List[int]) -> List[int]:
    """
    把 [0, -1, 2] 这种原始下标列表，转成合法正下标，去重后保序。
    """
    res: List[int] = []
    for i in raw_indices:
        j = i if i >= 0 else n + i  # 负下标转正
        if 0 <= j < n:
            res.append(j)

    out: List[int] = []
    seen = set()
    for j in res:
        if j not in seen:
            seen.add(j)
            out.append(j)
    return out


def get_values_by_key_path(item: Any, key_path: str) -> List[Any]:
    """
    输入一个 item 和 key_path，返回所有“走得到”的值（列表）。

    取值语义：

      .a.b.c       -> 取到所有 c 的值（通常是单个）
      .a.b[]       -> 取 a.b 这个列表里的所有元素
      .a.b[].c     -> 对 a.b 列表里的每个元素，取 c；不存在 c 的元素会被忽略
      .a.b[0,-1].c -> 对下标 0 和 -1 的元素，取 c；不存在 c 的元素会被忽略
    """
    steps = _parse_path(key_path)
    if not steps:
        # 空路径：直接返回 item 本身
        return [item]

    current_nodes: List[Any] = [item]

    for step in steps:
        next_nodes: List[Any] = []

        if step.mode == "dict":
            # 普通下钻：只在 dict 中取 key
            for node in current_nodes:
                if isinstance(node, dict) and step.key in node:
                    next_nodes.append(node[step.key])

        elif step.mode == "all":
            # key[]：把对应 list 展开
            for node in current_nodes:
                if isinstance(node, dict) and step.key in node:
                    lst = node[step.key]
                    if isinstance(lst, list):
                        next_nodes.extend(lst)

        elif step.mode == "indices":
            # key[0,-1]：只取指定下标的元素
            for node in current_nodes:
                if isinstance(node, dict) and step.key in node:
                    lst = node[step.key]
                    if isinstance(lst, list) and step.indices:
                        for j in _resolve_indices(len(lst), step.indices):
                            next_nodes.append(lst[j])

        current_nodes = next_nodes
        if not current_nodes:
            break

    return current_nodes


def read_jsonl(path_patterns, is_root=False):
    if isinstance(path_patterns, str):
        path_patterns = [path_patterns]
    for path_pattern in path_patterns:
        path_list = sorted(glob.glob(path_pattern, recursive=True))
        for path in path_list:
            with open(path, encoding="utf-8", errors="ignore") as reader:
                for line in reader:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        line_json = json.loads(line)
                        if isinstance(line_json, dict) and is_root:
                            line_json["##FILEPATH##"] = path
                        yield line_json
                    except Exception:
                        continue


def read_json(path_patterns, is_root=False):
    if isinstance(path_patterns, str):
        path_patterns = [path_patterns]
    for path_pattern in path_patterns:
        path_list = sorted(glob.glob(path_pattern, recursive=True))
        for path in path_list:
            with open(path, encoding="utf-8", errors="ignore") as reader:
                try:
                    line_json = json.load(reader)
                    if isinstance(line_json, dict) and is_root:
                        line_json["##FILEPATH##"] = path
                    yield line_json
                except Exception:
                    continue


def save_jsonl(samples, result_save_path):
    Path(result_save_path).parent.mkdir(parents=True, exist_ok=True)
    with open(result_save_path, "w", encoding="utf-8") as writer:
        for sample in samples:
            if isinstance(sample, dict):
                sample.pop("##FILEPATH##", None)
            writer.write(json.dumps(sample, ensure_ascii=False)+"\n")
            yield sample


def save_data_to_excel_merge(samples, file_path, key, merge=True):
    """
    生成表格，某些单元格要合并
    """
    samples = [x for x in samples]
    for s in samples:
        assert isinstance(s[key], list)
        assert len(s[key]) > 0
        assert isinstance(s[key][0], dict), s[key][0]
    column_names = []
    for k in samples[0]:
        if k != key:
            column_names.append(k)
        else:
            for k1 in samples[0][key][0]:
                column_names.append(f"{key}::{k1}")
    assert len(column_names) <= 26
    column_map = {y:x for x, y in zip(string.ascii_uppercase, column_names)}

    wb = openpyxl.Workbook()
    ws = wb.active
    # 标题行
    font_bold = Font(bold=True)
    for cn in column_names:
        ws[f"{column_map[cn]}1"] = cn
        ws[f"{column_map[cn]}1"].font = font_bold
    row_start = 2  # 第一行是标题
    for s in tqdm.tqdm(samples, "写入表格"):
        nrow = len(s[key])
        row_end = row_start + nrow - 1
        # merge单元格并填入内容
        for k in s:
            if k != key:
                if merge:
                    ws.merge_cells(f"{column_map[k]}{row_start}:{column_map[k]}{row_end}")
                try:
                    ws[f"{column_map[k]}{row_start}"] = s[k]
                except:
                    try:
                        ws[f"{column_map[k]}{row_start}"] = json.dumps(s[k], ensure_ascii=False)
                    except:
                        raise ValueError(s[k])
            else:
                for ii in range(len(s[key])):
                    for kk in s[key][ii]:
                        write_k = column_map[f"{k}::{kk}"]
                        try:
                            ws[f"{write_k}{row_start + ii}"] = s[key][ii][kk]
                        except:
                            try:
                                ws[f"{write_k}{row_start + ii}"] = json.dumps(s[key][ii][kk], ensure_ascii=False)
                            except:
                                raise ValueError(s[key][ii][kk])
        yield s
        row_start = row_start + nrow
    wb.save(file_path)
    print(f"✅ File saved: {file_path} - Records: {len(samples)}")
if __name__ == "__main__":
    import json

    obj = {
        "key1": {
            "key2": {
                "key3": [
                    {"key4": 123, "keep": "a"},
                    {"key4": 456, "keep": "b"},
                    {"key4": 789, "keep": "c"}
                ]
            }
        }
    }

    # # 1. 对所有元素删 key4
    # delete_field_by_path(obj, ".key1.key2.key3[].key4")

    # # 2. 删除key3
    # delete_field_by_path(obj, ".key1.key2.key3")
   
    # # 3. 删除key3所有元素
    # delete_field_by_path(obj, ".key1.key2.key3[]")

    # # 2. 只对第 0、1 个元素删 key4
    # delete_field_by_path(obj, ".key1.key2.key3[0,1].key4")

    # # 3. 直接删掉 key3 列表中第 0、2 个元素
    # delete_field_by_path(obj, ".key1.key2.key3[0,2]")

    # # 3. 直接删掉 key3 列表中第-1 个元素
    # delete_field_by_path(obj, ".key1.key2.key3[-3, -1]")

    # delete_field_by_path(obj, ".key1.key2.key3[-2,1].key4")

    # rename_field_by_path(obj, ".key1.key2.key3", "key333")

    # print(json.dumps(obj, ensure_ascii=False, indent=2))

    # data = {
    #     "a": {
    #         "b": [
    #             {"c": 1, "x": 0},
    #             {"c": 2, "y": 0},
    #         ]
    #     }
    # }

    # print(has_key_path(data, ".a.b.c"))         # False（b 不是 dict）
    # print(has_key_path(data, ".a.b[]"))         # True（b 是 list 且非空）
    # print(has_key_path(data, ".a.b[].c"))       # True（所有元素都有 c）
    # print(has_key_path(data, ".a.b[].x"))       # False（有元素没有 x）
    # print(has_key_path(data, ".a.b[].y"))       # False（有元素没有 y）
    # print(has_key_path(data, ".a.b[].z"))       # False（有元素没有 z）

    # print(has_key_path(data, ".a.b[0,-1].c"))   # True（0 和 -1 都有 c）
    # print(has_key_path(data, ".a.b[0,-1].y"))   # False（这两个元素并非都存在 y）
    # print(has_key_path(data, ".a.b[-1].y"))     # True（-1有y）
    # print(has_key_path(data, ".a.b[100].c"))    # False（下标越界）


    # item = {
    #     "a": {
    #         "b": [
    #             {"c": 1, "x": 0},
    #             {"c": 2, "x": 0},
    #             {"d": 3, "x": 0}
    #         ]
    #     }
    # }

    # print(get_values_by_key_path(item, ".a.b.c"))
    # # []   因为 b 是 list，不是 dict，.b.c 走不下去

    # print(get_values_by_key_path(item, ".a.b[]"))
    # # [{'c': 1, 'x': 0}, {'c': 2, 'x': 0}, {'d': 3, 'x': 0}]

    # print(get_values_by_key_path(item, ".a.b"))  # NOTE 注意这个和上面的结果不一样
    # # [[{'c': 1, 'x': 0}, {'c': 2, 'x': 0}, {'d': 3, 'x': 0}]]

    # print(get_values_by_key_path(item, ".a.b[].c"))
    # # [1, 2]   第三个元素没有 c，被自动忽略

    # print(get_values_by_key_path(item, ".a.b[0,-1].c"))
    # # [1]      下标 0 有 c，下标 -1 没 c，所以只收集到 1
